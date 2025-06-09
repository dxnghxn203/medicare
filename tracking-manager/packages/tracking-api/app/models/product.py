import asyncio
import re
import unicodedata
import json
from io import BytesIO
from types import SimpleNamespace
import platform
import socket
import time
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from types import SimpleNamespace
import zipfile
import os
import olefile
import pandas as pd
from fastapi import UploadFile
from pydantic import ValidationError
from pymongo import UpdateOne, ReplaceOne
from starlette import status
from typing import Set, Dict, Tuple
from datetime import datetime, timedelta

from starlette.responses import JSONResponse

from app.core import logger, response, recommendation
from app.core.database import db
from app.core.s3 import upload_file, upload_any_file
from app.entities.pharmacist.response import ItemPharmacistRes
from app.entities.product.request import ItemProductDBInReq, ItemImageDBReq, ItemPriceDBReq, ItemProductDBReq, \
    UpdateCategoryReq, ItemCategoryDBReq, ApproveProductReq, UpdateProductStatusReq, \
    ItemUpdateProductReq, ItemIngredientDBReq, ItemManufacturerDBReq, ItemProductImportReq, ItemProductInventoryReq
from app.entities.product.response import ItemProductDBRes, ItemProductImportRes, ItemProductInventoryRes
from app.helpers import redis
from app.helpers.constant import generate_id
from app.helpers.time_utils import get_current_time
from app.models.category import get_all_categories, get_all_categories_admin
from app.models.comment import count_comments
from app.models.review import count_reviews, average_rating

collection_name = "products"
collection_category = "categories"
collection_import = "products_imports"
collection_inventory = "products_inventory"

VALID_MEDIA_TYPES = {"images", "images_primary", "certificate_file"}

async def get_product_by_slug(slug: str):
    try:
        collection = db[collection_name]
        cur = collection.find_one({
            "slug": slug,
            "is_approved": True,
            "active": True
        })
        if cur:
            count_review = await count_reviews(cur["product_id"])
            count_comment = await count_comments(cur["product_id"])
            avg_rating = await average_rating(cur["product_id"])
            return ItemProductDBRes(**cur,
                                   count_review=count_review,
                                   count_comment=count_comment,
                                   rating=avg_rating)
        return None
    except Exception as e:
        raise e

def build_match_conditions(main_category: str = None, low_stock_status: bool = None):
    match_conditions = []

    if main_category:
        match_conditions.append({
            "$match": {
                "category.main_category_id": main_category
            }
        })

    if low_stock_status is not None:
        threshold = 10 if low_stock_status else 0
        stock_expr = {
            "$cond": {
                "if": {
                    "$lt" if low_stock_status else "$lte": [
                        {"$subtract": ["$$item.inventory", "$$item.sell"]},
                        threshold
                    ]
                },
                "then": True,
                "else": False
            }
        }

        match_conditions.append({
            "$match": {
                "$expr": {
                    "$gt": [
                        {
                            "$size": {
                                "$filter": {
                                    "input": "$prices",
                                    "as": "item",
                                    "cond": stock_expr
                                }
                            }
                        },
                        0
                    ]
                }
            }
        })

    return match_conditions

async def get_all_product(
    page: int, page_size: int,
    low_stock_status: bool = None,
    main_category: str = None,
    best_seller: bool = None
):
    try:
        collection = db[collection_name]
        skip_count = (page - 1) * page_size

        if best_seller:
            base_match = build_match_conditions(main_category, low_stock_status)

            pipeline = base_match + [
                {
                    "$addFields": {
                        "total_sell": {
                            "$sum": {
                                "$map": {
                                    "input": "$prices",
                                    "as": "p",
                                    "in": {
                                        "$multiply": ["$$p.sell", "$$p.amount"]
                                    }
                                }
                            }
                        }
                    }
                },
                {
                    "$sort": {"total_sell": -1, "created_at": -1}
                },
                {"$skip": skip_count},
                {"$limit": page_size}
            ]

            pipeline_total = base_match + [{"$count": "total"}]

            product_list = collection.aggregate(pipeline).to_list(length=page_size)
            total_result = collection.aggregate(pipeline_total).to_list(length=1)
            total = total_result[0]["total"] if total_result else 0

        else:
            query = {}

            if main_category:
                query["category.main_category_id"] = main_category

            if low_stock_status is not None:
                threshold = 10 if low_stock_status else 0
                stock_expr = {
                    "$cond": {
                        "if": {
                            "$lt" if low_stock_status else "$lte": [
                                {"$subtract": ["$$item.inventory", "$$item.sell"]},
                                threshold
                            ]
                        },
                        "then": True,
                        "else": False
                    }
                }
                query["$expr"] = {
                    "$gt": [
                        {
                            "$size": {
                                "$filter": {
                                    "input": "$prices",
                                    "as": "item",
                                    "cond": stock_expr
                                }
                            }
                        },
                        0
                    ]
                }

            product_list = collection.find(query).sort("created_at", -1).skip(skip_count).limit(page_size).to_list(length=page_size)
            total = collection.count_documents(query)

        return {
            "total_products": total,
            "products": [ItemProductDBRes.from_mongo(p) for p in product_list]
        }

    except Exception as e:
        logger.error(f"Failed [get_all_product]: {e}")
        raise e

async def get_discount_product(page: int, page_size: int):
    try:
        collection = db[collection_name]
        skip_count = (page - 1) * page_size
        now = get_current_time()

        product_list = collection.find({
            "is_approved": True,
            "active": True,
            "prices": {
                "$elemMatch": {
                    "discount": {"$gt": 0},
                    "expired_date": {"$gt": now}
                }
            }
        }).skip(skip_count).limit(page_size)

        result = []
        for product in product_list:
            product["prices"] = [
                p for p in product.get("prices", [])
                if p.get("discount", 0) > 0 and
                   p.get("expired_date") and
                p["expired_date"] > now
            ]
            product_id = product["product_id"]
            count_review, count_comment, avg_rating = await asyncio.gather(
                count_reviews(product_id),
                count_comments(product_id),
                average_rating(product_id)
            )

            product["count_review"] = count_review
            product["count_comment"] = count_comment
            product["rating"] = avg_rating

            result.append(ItemProductDBRes(**product))

        return result
    except Exception as e:
        logger.error(f"Failed [get_discount_product]: {e}")
        raise e

async def get_discount_product_for_admin(page: int, page_size: int, is_approved: bool):
    try:
        collection = db[collection_name]
        skip_count = (page - 1) * page_size
        now = get_current_time()

        query = {
            "is_approved": is_approved,
            "prices": {
                "$elemMatch": {
                    "discount": {"$gt": 0},
                    "expired_date": {"$gt": now}
                }
            }
        }

        product_list = collection.find(query).skip(skip_count).limit(page_size)

        return {
            "total_products": collection.count_documents(query),
            "products": [ItemProductDBRes(**product) for product in product_list]
        }
    except Exception as e:
        logger.error(f"Failed [get_discount_product_for_admin]: {e}")
        raise e

async def get_product_top_selling(top_n):
    try:
        result = recommendation.send_request("/v1/top-selling/", {"top_n": top_n})
        product_list = result["data"]

        enriched_products = []

        for product in product_list:
            product_id = product["product_id"]

            count_review, count_comment, avg_rating = await asyncio.gather(
                count_reviews(product_id),
                count_comments(product_id),
                average_rating(product_id)
            )

            product["count_review"] = count_review
            product["count_comment"] = count_comment
            product["rating"] = avg_rating

            enriched_products.append(product)

        result["data"] = enriched_products
        return result
    except Exception as e:
        logger.error(f"Failed [get_product_top_selling]: {e}")
        return response.BaseResponse(
            status="failed",
            message="Internal server error",
        )

async def get_related_product(product_id, top_n=5):
    try:
        result = recommendation.send_request("/v1/related/", {"product_id": product_id, "top_n": top_n})
        product_list = result["data"]
        if not product_list:
            result["data"] = []
            return result

        enriched_products = []

        for product in product_list:
            product_id = product["product_id"]

            count_review, count_comment, avg_rating = await asyncio.gather(
                count_reviews(product_id),
                count_comments(product_id),
                average_rating(product_id)
            )

            product["count_review"] = count_review
            product["count_comment"] = count_comment
            product["rating"] = avg_rating

            enriched_products.append(product)

        result["data"] = enriched_products
        return result
    except Exception as e:
        logger.error(f"Failed [get_related_product]: {e}")
        raise e

async def get_product_by_cart_mongo(product_ids, cart):
    try:
        collection = db[collection_name]
        products = list(collection.find({"product_id": {"$in": product_ids}}, {"_id": 0}))
        detailed_cart = []
        for product in products:
            product_id = product["product_id"]
            cart_item = cart.get(product_id)
            if cart_item:
                price_id = cart_item["price_id"]
                quantity = cart_item["quantity"]

                matching_price = next((p for p in product["prices"] if p["price_id"] == price_id), None)
                if matching_price:
                    detailed_cart.append({"product": ItemProductDBRes(**product), "price_id": price_id, "quantity": quantity})

        return detailed_cart

    except Exception as e:
        logger.error(f"Failed [get_product_by_cart_mongo]: {e}")
        return []

async def get_product_by_cart_id(product_ids, cart):
    try:
        collection = db[collection_name]
        products = list(collection.find({
            "product_id": {"$in": product_ids},
            "is_approved": True,
            "active": True
        }, {"_id": 0}))
        detailed_cart = []
        for cart_item in cart:
            product_id = cart_item["product_id"]
            price_id = cart_item["price_id"]
            quantity = cart_item["quantity"]

            product = next((p for p in products if p["product_id"] == product_id), None)
            if not product:
                continue

            matching_price = next((p for p in product.get("prices", []) if p["price_id"] == price_id), None)
            if matching_price:
                detailed_cart.append({
                    "product": ItemProductDBRes(**product),
                    "price_id": price_id,
                    "quantity": quantity
                })

        return detailed_cart

    except Exception as e:
        logger.error(f"Failed [get_product_by_cart_id]: {e}")
        return []

async def get_product_by_list_id(product_ids):
    try:
        collection = db[collection_name]
        product_list = collection.find({
            "product_id": {"$in": product_ids},
            "is_approved": True,
            "active": True
        })
        enriched_products = []

        for prod in product_list:
            product_id = prod["product_id"]
            count_review, count_comment, avg_rating = await asyncio.gather(
                count_reviews(product_id),
                count_comments(product_id),
                average_rating(product_id),
            )

            enriched_products.append(
                ItemProductDBRes(
                    **prod,
                    count_review=count_review,
                    count_comment=count_comment,
                    rating=avg_rating
                )
            )
        return enriched_products
    except Exception as e:
        logger.error(f"Failed [get_product_by_list_id]: {e}")
        return []

async def get_product_featured(main_category_id, sub_category_id=None, child_category_id=None,  top_n=5):
    try:
        params = {
            "main_category_id": main_category_id,
            "sub_category_id": sub_category_id,
            "child_category_id": child_category_id,
            "top_n": top_n
        }
        filtered_params = {k: v for k, v in params.items() if v is not None}

        result = recommendation.send_request("/v1/featured/", filtered_params)
        product_list = result["data"]

        enriched_products = []

        for product in product_list:
            product_id = product["product_id"]

            count_review, count_comment, avg_rating = await asyncio.gather(
                count_reviews(product_id),
                count_comments(product_id),
                average_rating(product_id)
            )

            product["count_review"] = count_review
            product["count_comment"] = count_comment
            product["rating"] = avg_rating

            enriched_products.append(product)

        result["data"] = enriched_products
        return result

    except Exception as e:
        logger.error(f"Failed [get_featured]: {e}")
        return response.BaseResponse(
            status="failed",
            message="Internal server error",
        )

async def add_product_db(item: ItemProductDBInReq, images_primary, images, email, certificate_file=None):
    try:
        product_collection = db[collection_name]
        inventory_collection = db[collection_inventory]
        category_collection = db[collection_category]

        cur = product_collection.find_one({"slug": item.slug})
        if cur:
            raise response.JsonException(
                status_code=status.HTTP_400_BAD_REQUEST,
                message=f"Sản phẩm đã tồn tại: {item.slug}"
            )

        certificate_url = (upload_any_file(certificate_file, "certificates") or "") if certificate_file else ""

        image_list = []
        if images:
            image_list = [
                ItemImageDBReq(
                    images_id=generate_id("IMAGE"),
                    images_url=file_url,
                ) for idx, img in enumerate(images or []) if (file_url := upload_file(img, "images"))
            ]
        logger.info(f"{image_list}")

        price_list = []
        inventory_docs = []
        product_id = generate_id("PRODUCT")

        if item.prices and item.prices.prices:
            for price in item.prices.prices:
                price_id = generate_id("PRICE")
                price_obj = ItemPriceDBReq(
                    price_id=price_id,
                    price=price.original_price * (100 - price.discount) / 100,
                    discount=price.discount,
                    unit=price.unit,
                    weight=price.weight,
                    amount=price.amount,
                    original_price=price.original_price,
                    inventory=price.inventory,
                    expired_date=price.expired_date
                )
                price_list.append(price_obj)

                inventory_obj = ItemProductInventoryReq(
                    product_id=product_id,
                    price_id=price_id,
                    amount=price.amount,
                    inventory=price.inventory,
                )

                inventory_docs.append(inventory_obj.dict())

        ingredients_list = item.ingredients.ingredients if item.ingredients and item.ingredients.ingredients else []
        images_primary_url = (upload_file(images_primary, "images_primary") or "") if images_primary else ""

        main_category = category_collection.find_one({"main_category_id": item.category.main_category_id})
        if not main_category:
            raise response.JsonException(
                status_code=status.HTTP_400_BAD_REQUEST,
                message="Danh mục chính không hợp lệ"
            )

        sub_category = next(
            (sub for sub in main_category.get("sub_category", []) if sub["sub_category_id"] == item.category.sub_category_id),
            None
        )
        if not sub_category:
            raise response.JsonException(
                status_code=status.HTTP_400_BAD_REQUEST,
                message="Danh mục phụ không hợp lệ"
            )

        child_category = next(
            (child for child in sub_category.get("child_category", []) if
             child["child_category_id"] == item.category.child_category_id),
            None
        )
        if not child_category:
            raise response.JsonException(
                status_code=status.HTTP_400_BAD_REQUEST,
                message="Danh mục con không hợp lệ"
            )

        category_obj = ItemCategoryDBReq(
            main_category_id=item.category.main_category_id,
            main_category_name=main_category.get("main_category_name", ""),
            main_category_slug=main_category.get("main_category_slug", ""),
            sub_category_id=item.category.sub_category_id,
            sub_category_name=sub_category.get("sub_category_name", ""),
            sub_category_slug=sub_category.get("sub_category_slug", ""),
            child_category_id=item.category.child_category_id,
            child_category_name=child_category.get("child_category_name", ""),
            child_category_slug=child_category.get("child_category_slug", ""),
        )

        item_data = ItemProductDBReq(
            **{k: v for k, v in dict(item).items() if k not in ["prices", "ingredients", "category"]},
            product_id=product_id,
            prices=price_list,
            images=image_list,
            ingredients=ingredients_list,
            images_primary=images_primary_url,
            category=category_obj,
            certificate_file=certificate_url,
            created_by=email,
            updated_by=email
        )

        insert_result = product_collection.insert_one(item_data.dict())
        logger.info(f"Thêm sản phẩm thành công: {insert_result.inserted_id}")

        if inventory_docs:
            inventory_collection.insert_many(inventory_docs)
            logger.info(f"Đã thêm {len(inventory_docs)} bản ghi vào bảng products_inventory")

    except Exception as e:
        logger.error(f"Lỗi khi thêm sản phẩm: {e}")
        raise e

async def update_product_category(item: UpdateCategoryReq, email: str):
    try:
        collection = db[collection_name]
        category_collection = db[collection_category]
        product = collection.find_one({"product_id": item.product_id})

        if not product:
            raise response.JsonException(
                status_code=status.HTTP_404_NOT_FOUND,
                message="Không tìm thấy sản phẩm"
            )

        main_category = category_collection.find_one({"main_category_id": item.main_category_id})
        if not main_category:
            raise response.JsonException(
                status_code=status.HTTP_400_BAD_REQUEST,
                message="Danh mục chính không hợp lệ"
            )

        sub_category = next(
            (sub for sub in main_category.get("sub_category", []) if sub["sub_category_id"] == item.sub_category_id),
            None
        )
        if not sub_category:
            raise response.JsonException(
                status_code=status.HTTP_400_BAD_REQUEST,
                message="Danh mục phụ không hợp lệ"
            )

        child_category = next(
            (child for child in sub_category.get("child_category", []) if
             child["child_category_id"] == item.child_category_id),
            None
        )
        if not child_category:
            raise response.JsonException(
                status_code=status.HTTP_400_BAD_REQUEST,
                message="Danh mục con không hợp lệ"
            )

        update_data = {
            "category.main_category_id": item.main_category_id,
            "category.main_category_name": main_category.get("main_category_name", ""),
            "category.main_category_slug": main_category.get("main_category_slug", ""),

            "category.sub_category_id": item.sub_category_id,
            "category.sub_category_name": sub_category.get("sub_category_name", ""),
            "category.sub_category_slug": sub_category.get("sub_category_slug", ""),

            "category.child_category_id": item.child_category_id,
            "category.child_category_name": child_category.get("child_category_name", ""),
            "category.child_category_slug": child_category.get("child_category_slug", ""),

            "updated_by": email,
            "updated_at": get_current_time()
        }

        collection.update_one({"product_id": item.product_id}, {"$set": update_data})
        logger.info(f"Updated category names for product_id: {item.product_id}")

        return response.SuccessResponse(message="Cập nhật danh mục thành công")
    except Exception as e:
        logger.error(f"Error updating product category: {str(e)}")
        raise e

async def delete_product(product_id: str):
    try:
        collection = db[collection_name]
        product = collection.find_one({"product_id": product_id})
        if not product:
            logger.info(f"Product not found for product_id: {product_id}")
        else:
            product = ItemProductDBRes(**product)
            delete_result = collection.delete_one({"product_id": product.product_id})
            if delete_result.deleted_count == 0:
                logger.info(f"Product not deleted for product_id: {product_id}")

            inventory_collection = db[collection_inventory]
            delete_inventory = inventory_collection.delete_many({"product_id": product_id})
            if delete_inventory.deleted_count == 0:
                logger.info(f"Inventory not deleted for product_id: {product_id}")

        return response.SuccessResponse(message="Xóa sản phẩm thành công")
    except Exception as e:
        logger.error(f"Error deleting product: {str(e)}")
        raise e

async def getAvailableQuantity(product_id: str, price_id: str):
    try:
        product = await get_product_by_id(product_id, price_id)

        if not product:
            raise response.JsonException(
                status_code=status.HTTP_404_NOT_FOUND,
                message="Product not found"
            )
        sell = product.prices[0].sell
        inventory = product.prices[0].inventory

        logger.info(f"Sell: {sell}, Inventory: {inventory}")

        return inventory - sell

    except Exception as e:
        logger.error(f"Error getting available quantity: {str(e)}")
        return None

async def get_product_by_id(product_id: str, price_id: str, is_admin: bool = False):
    try:
        collection = db[collection_name]

        query = {
            "product_id": product_id,
            "prices": {
                "$elemMatch": {"price_id": price_id}
            }
        }
        if not is_admin:
            query.update({
                "is_approved": True,
                "active": True
            })

        product = collection.find_one(query, {"_id": 0})

        if not product:
            logger.info(f"Product not found for product_id: {product_id}")
            raise response.JsonException(
                status_code=status.HTTP_404_NOT_FOUND,
                message="Product not found"
            )

        product["prices"] = [
            price for price in product.get("prices", [])
            if price.get("price_id") == price_id
        ]

        return ItemProductDBRes(**product)
    except Exception as e:
        logger.error(f"Error getting product by id: {str(e)}")
        raise e

async def get_product_inventory(product_id: str, price_id: str):
    try:
        inventory_collection = db["products_inventory"]
        inventory_doc = inventory_collection.find_one({
            "product_id": product_id,
            "price_id": price_id
        })

        if not inventory_doc:
            raise response.JsonException(
                status_code=status.HTTP_404_NOT_FOUND,
                message="Không tìm thấy thông tin tồn kho"
            )

        return ItemProductInventoryRes(**inventory_doc)

    except Exception as e:
        logger.error(f"Error getting product inventory: {str(e)}")
        raise e

async def restore_product_sell(product_id: str, price_id: str, quantity: int):
    try:
        collection = db[collection_name]
        inventory_collection = db[collection_inventory]

        result = collection.update_one(
            {"product_id": product_id, "prices.price_id": price_id},
            {"$inc": {"prices.$.sell": -quantity}}
        )

        inventory_result = inventory_collection.update_one(
            {"product_id": product_id, "price_id": price_id},
            {"$inc": {"sell": -quantity}}
        )

        if result.modified_count == 0:
            raise response.JsonException(
                status_code=status.HTTP_404_NOT_FOUND,
                message="Product not found or quantity not updated"
            )

        if inventory_result.modified_count == 0:
            raise response.JsonException(
                status_code=status.HTTP_404_NOT_FOUND,
                message="Inventory not found or quantity not updated"
            )

        logger.info(f"Product sell restored successfully for product_id: {product_id}, price_id: {price_id}")
    except Exception as e:
        logger.error(f"Error restoring product sell: {str(e)}")
        raise e

async def get_product_best_deals(top_n: int):
    try:
        result = recommendation.send_request("/v1/top-selling/", {"top_n": top_n})
        product_list = result.get("data", [])

        enriched_products = []

        for product in product_list:
            product_id = product["product_id"]

            count_review, count_comment, avg_rating = await asyncio.gather(
                count_reviews(product_id),
                count_comments(product_id),
                average_rating(product_id)
            )

            prices = product.get("prices", [])
            max_discount = max([p.get("discount", 0) for p in prices]) if prices else 0

            product.update({
                "count_review": count_review,
                "count_comment": count_comment,
                "rating": avg_rating,
                "max_discount_percent": max_discount
            })

            enriched_products.append(product)

        sorted_products = sorted(
            enriched_products,
            key=lambda x: (
                -x.get("max_discount_percent", 0)
                -(x.get("rating") or 0),
            )
        )

        return response.BaseResponse(
            status="success",
            message="Best deal products retrieved successfully",
            data=sorted_products[:top_n]
        )

    except Exception as e:
        logger.error(f"Failed [get_product_best_deals]: {e}")
        return response.BaseResponse(
            status="failed",
            message="Internal server error"
        )

async def approve_product(item: ApproveProductReq, pharmacist: ItemPharmacistRes):
    try:
        collection = db[collection_name]
        product = collection.find_one({"product_id": item.product_id})
        if not product:
            raise response.JsonException(
                status_code=status.HTTP_404_NOT_FOUND,
                message="Product not found"
            )
        product_info = ItemProductDBRes(**product)
        if product_info.is_approved:
            raise response.JsonException(
                status_code=status.HTTP_400_BAD_REQUEST,
                message="Sản phẩm dã duyệt"
            )
        if product_info.verified_by and product_info.verified_by != pharmacist.email:
            raise response.JsonException(
                status_code=status.HTTP_400_BAD_REQUEST,
                message="Không có quyền duyệt sản phẩm này"
            )
        collection.update_one({
            "product_id": item.product_id},
            {"$set": {
                "is_approved": item.is_approved,
                "verified_by": pharmacist.email,
                "rejected_note": item.rejected_note,
                "pharmacist_name": pharmacist.user_name,
                "pharmacist_gender": pharmacist.gender,
            }
        })
        logger.info(f"Product approved successfully for {item.product_id} by {pharmacist.email} with: {item.is_approved}")

        return response.SuccessResponse(message="Product approved successfully")
    except Exception as e:
        logger.error(f"Error approving product: {str(e)}")
        raise e

async def get_approved_product(
        email: str, page: int = 1, page_size: int = 10,
        keyword: str = None,
        main_category: str = None,
        prescription_required: bool = None,
        status: str = None  # "approved", "pending", or None for all
):
    try:
        collection = db[collection_name]
        skip_count = (page - 1) * page_size

        query = {}

        if keyword:
            pattern = re.escape(keyword)
            query["name_primary"] = {"$regex": pattern, "$options": "i"}

        if status == "approved":
            query["is_approved"] = True
            query["verified_by"] = email
        elif status == "pending":
            query["is_approved"] = False
        else:
            query["verified_by"] = {"$in": [None, "", email]}

        if main_category:
            query["category.main_category_id"] = main_category

        if prescription_required is not None:
            query["prescription_required"] = prescription_required

        logger.info(f"Query: {query}")

        product_list = collection.find(query).sort("created_at", -1).skip(skip_count).limit(page_size).to_list(length=page_size)
        total = collection.count_documents(query)

        return {
            "total_products": total,
            "products": [ItemProductDBRes.from_mongo(p) for p in product_list]
        }
    except Exception as e:
        logger.error(f"Failed [get_approved_product]: {e}")
        raise e

async def update_product_status(item: UpdateProductStatusReq, email: str):
    try:
        collection = db[collection_name]
        collection.update_one({
                "product_id": item.product_id
            },
            {
                "$set": {
                    "active": item.status,
                    "updated_by": email,
                    "updated_at": get_current_time()
                }
            }
        )
        return response.SuccessResponse(message="Product status updated successfully")
    except Exception as e:
        logger.error(f"Error updating product status: {str(e)}")
        raise e

async def update_product_images_primary(product_id: str, file, email: str):
    try:
        if not file:
            raise response.JsonException(status_code=status.HTTP_400_BAD_REQUEST, message="File không hợp lệ")

        collection = db[collection_name]
        product = collection.find_one({"product_id": product_id})
        if not product:
            raise response.JsonException(status_code=status.HTTP_400_BAD_REQUEST, message="Không tìm thấy sản phẩm")

        url = upload_file(file, "images_primary")
        collection.update_one(
            {"product_id": product_id},
            {"$set": {"images_primary": url, "updated_by": email, "updated_at": get_current_time()}},
        )
        return response.SuccessResponse(message="Cập nhật ảnh chính thành công")
    except Exception as e:
        logger.error(f"Error updating product images primary: {str(e)}")
        raise e

async def update_product_certificate_file(product_id: str, file, email: str):
    try:
        if not file:
            raise response.JsonException(status_code=status.HTTP_400_BAD_REQUEST, message="File không hợp lệ")

        collection = db[collection_name]
        product = collection.find_one({"product_id": product_id})
        if not product:
            raise response.JsonException(status_code=status.HTTP_400_BAD_REQUEST, message="Không tìm thấy sản phẩm")

        url = upload_any_file(file, "certificates")
        collection.update_one(
            {"product_id": product_id},
            {"$set": {"certificate_file": url, "updated_by": email, "updated_at": get_current_time()}},
        )
        return response.SuccessResponse(message="Cập nhật chứng nhận thành công")
    except Exception as e:
        logger.error(f"Error updating product certificate file: {str(e)}")
        raise e

async def update_product_images(product_id: str, files, email: str):
    try:
        if not files:
            raise response.JsonException(status_code=status.HTTP_400_BAD_REQUEST, message="File không hợp lệ")

        collection = db[collection_name]
        product = collection.find_one({"product_id": product_id})
        if not product:
            raise response.JsonException(status_code=status.HTTP_400_BAD_REQUEST, message="Không tìm thấy sản phẩm")

        new_images = []
        for file in files:
            url = upload_file(file, "images")
            if url:
                new_images.append({
                    "images_id": generate_id("IMAGE"),
                    "images_url": url
                })
                time.sleep(0.001)

        collection.update_one(
            {"product_id": product_id},
            {"$set": {"images": new_images, "updated_by": email, "updated_at": get_current_time()}},
        )
        return response.SuccessResponse(message="Cập nhật images thành công")
    except Exception as e:
        logger.error(f"Error updating product images: {str(e)}")
        raise e

async def update_product_fields(update_data: ItemUpdateProductReq, email):
    try:
        collection = db[collection_name]
        inventory_collection = db[collection_inventory]
        product = collection.find_one({"product_id": update_data.product_id})
        if not product:

            raise response.JsonException(
                status_code=status.HTTP_400_BAD_REQUEST,
                message="Không tìm thấy sản phẩm"
            )

        cur = collection.find_one({
            "slug": update_data.slug,
            "product_id": {"$ne": update_data.product_id}
        })
        if cur:
            raise response.JsonException(
                status_code=status.HTTP_400_BAD_REQUEST,
                message=f"Sản phẩm đã tồn tại với slug: {update_data.slug}"
            )

        update_fields = {}

        for field in [
            "product_name", "name_primary", "slug", "description", "origin",
            "uses", "dosage", "side_effects", "precautions", "storage", "dosage_form", "brand",
            "prescription_required", "registration_number", "full_descriptions"
        ]:
            value = getattr(update_data, field, None)
            if value is not None:
                update_fields[field] = value

        if update_data.prices:
            price_list = []
            for price in update_data.prices:
                price_info = await get_product_by_id(update_data.product_id, price.price_id, is_admin=True)
                price_obj = ItemPriceDBReq(
                    price_id=price.price_id,
                    price=price.original_price * (100 - price.discount) / 100,
                    discount=price.discount,
                    unit=price.unit,
                    weight=price.weight,
                    amount=price.amount,
                    original_price=price.original_price,
                    inventory=price.inventory,
                    sell=price_info.prices[0].sell,
                    delivery=price_info.prices[0].delivery,
                    expired_date=price.expired_date
                )
                price_list.append(price_obj.dict())

                inventory_collection.update_one(
                    {"product_id": update_data.product_id, "price_id": price.price_id},
                    {"$set": {
                        "amount": price.amount,
                        "inventory": price.inventory
                    }}
                )
            update_fields["prices"] = price_list

        if update_data.ingredients and update_data.ingredients.ingredients:
            update_fields["ingredients"] = [i.dict() for i in update_data.ingredients.ingredients]

        if update_data.manufacturer:
            update_fields["manufacturer"] = update_data.manufacturer.dict()

        if update_data.category:
            category_collection = db[collection_category]
            main_category = category_collection.find_one({"main_category_id": update_data.category.main_category_id})
            if not main_category:
                raise response.JsonException(status_code=status.HTTP_400_BAD_REQUEST, message="Danh mục chính không hợp lệ")

            sub_category = next(
                (sub for sub in main_category.get("sub_category", []) if sub["sub_category_id"] == update_data.category.sub_category_id),
                None
            )
            if not sub_category:
                raise response.JsonException(status_code=status.HTTP_400_BAD_REQUEST, message="Danh mục phụ không hợp lệ")

            child_category = next(
                (child for child in sub_category.get("child_category", []) if child["child_category_id"] == update_data.category.child_category_id),
                None
            )
            if not child_category:
                raise response.JsonException(status_code=status.HTTP_400_BAD_REQUEST, message="Danh mục con không hợp lệ")

            category_obj = ItemCategoryDBReq(
                main_category_id=update_data.category.main_category_id,
                main_category_name=main_category.get("main_category_name", ""),
                main_category_slug=main_category.get("main_category_slug", ""),
                sub_category_id=update_data.category.sub_category_id,
                sub_category_name=sub_category.get("sub_category_name", ""),
                sub_category_slug=sub_category.get("sub_category_slug", ""),
                child_category_id=update_data.category.child_category_id,
                child_category_name=child_category.get("child_category_name", ""),
                child_category_slug=child_category.get("child_category_slug", "")
            )
            update_fields["category"] = category_obj.dict()

        if update_fields:
            update_fields.update({
                "is_approved": False,
                "verified_by": "",
                "rejected_note": "",
                "pharmacist_name": "",
                "pharmacist_gender": "",
                "updated_by": email,
                "updated_at": get_current_time()
            })
            collection.update_one({"product_id": update_data.product_id}, {"$set": update_fields})

        return response.SuccessResponse(message="Cập nhật sản phẩm thành công")
    except Exception as e:
        logger.error(f"Error updating product fields: {str(e)}")
        raise e

async def search_products_by_name(keyword: str):
    try:
        collection = db[collection_name]
        pattern = re.escape(keyword)
        query = {
            "name_primary": {"$regex": pattern, "$options": "i"},
            "is_approved": True,
            "active": True
        }
        product_list = collection.find(query)

        enriched_products = []

        for prod in product_list:
            product_id = prod["product_id"]
            count_review, count_comment, avg_rating = await asyncio.gather(
                count_reviews(product_id),
                count_comments(product_id),
                average_rating(product_id),
            )

            enriched_products.append(
                ItemProductDBRes(
                    **prod,
                    count_review=count_review,
                    count_comment=count_comment,
                    rating=avg_rating
                )
            )
        return enriched_products
    except Exception as e:
        logger.error(f"Error searching products by name: {str(e)}")
        raise e

async def get_product_brands():
    try:
        collection = db[collection_name]
        product_list = collection.distinct(
            "brand",
            {
                "brand": {"$ne": None},
                "is_approved": True,
                "active": True
            }
        )
        return product_list
    except Exception as e:
        logger.error(f"Error searching products by name: {str(e)}")
        raise e

def normalize_text(text):
    if not text:
        return ""
    text = unicodedata.normalize("NFD", text)
    text = text.encode("ascii", "ignore").decode("utf-8")
    text = re.sub(r"[\s\-_]+", "", text).lower()
    return text

def normalize_manufacturer_contact(contact_value):
    contact_str = str(contact_value).strip()
    if contact_str.isdigit() and not contact_str.startswith("0"):
        contact_str = "0" + contact_str
    return contact_str

async def extract_category_object(row: dict, all_categories: list) -> Tuple[ItemCategoryDBReq, str]:
    try:
        errors = []
        normalized_main = normalize_text(row.get("main_category_name", ""))
        normalized_sub = normalize_text(row.get("sub_category_name", ""))
        normalized_child = normalize_text(row.get("child_category_name", ""))

        matched_main = next((cat for cat in all_categories if normalize_text(cat.get("main_category_name", "")) == normalized_main), None)
        if not matched_main and all_categories:
            matched_main = all_categories[0]
            errors.append(
                f"Danh mục chính không hợp lệ: {row.get('main_category_name')}, dùng mặc định '{matched_main.get('main_category_name')}'")
        elif not matched_main:
            return None, "Không tìm thấy danh mục chính nào"

        matched_sub = next((sub for sub in matched_main.get("sub_category", []) if normalize_text(sub.get("sub_category_name", "")) == normalized_sub), None)
        if not matched_sub and matched_main.get("sub_category"):
            matched_sub = matched_main.get("sub_category")[0]
            errors.append(
                f"Danh mục phụ không hợp lệ: {row.get('sub_category_name')}, dùng mặc định '{matched_sub.get('sub_category_name')}'")
        elif not matched_sub:
            return None, f"Danh mục phụ không xác định và không có danh sách danh mục phụ cho '{matched_main.get('main_category_name')}'"

        matched_child = next((child for child in matched_sub.get("child_category", []) if normalize_text(child.get("child_category_name", "")) == normalized_child), None)
        if not matched_child and matched_sub.get("child_category"):
            matched_child = matched_sub.get("child_category")[0]
            errors.append(
                f"Danh mục con không hợp lệ: {row.get('child_category_name')}, dùng mặc định '{matched_child.get('child_category_name')}'")
        elif not matched_child:
            return None, f"Danh mục con không xác định và không có danh sách danh mục con cho '{matched_sub.get('sub_category_name')}'"

        return ItemCategoryDBReq(
            main_category_id=matched_main["main_category_id"],
            main_category_name=matched_main["main_category_name"],
            main_category_slug=matched_main["main_category_slug"],
            sub_category_id=matched_sub["sub_category_id"],
            sub_category_name=matched_sub["sub_category_name"],
            sub_category_slug=matched_sub["sub_category_slug"],
            child_category_id=matched_child["child_category_id"],
            child_category_name=matched_child["child_category_name"],
            child_category_slug=matched_child["child_category_slug"]
        ), "; ".join(errors) if errors else None
    except Exception as e:
        logger.error(f"Error extracting category object: {str(e)}")
        return None, str(e)

async def extract_images_direct(sheet, df, row_idx, image_columns, is_primary=False):
    logger.info(
        f"[Image Extract] Đang chạy trên hệ điều hành: {platform.platform()} | Hostname: {socket.gethostname()}")
    image_results = {}  # key: column name, value: (url or None, error or None)
    img_row = 0

    logger.info(f"[Image Check] sheet._images = {sheet._images}")
    logger.info(f"[Image Count] Tổng số ảnh trong sheet: {len(sheet._images)}")

    for image in sheet._images:
        try:
            anchor = image.anchor._from
            img_row = anchor.row + 1
            img_col = anchor.col + 1
            img_col_letter = get_column_letter(img_col)
            header_value = sheet[f"{img_col_letter}1"].value

            logger.info(f"[Image Extract] Đang xử lý ảnh tại hàng {img_row}, cột {img_col} ({header_value})")

            if img_row == row_idx + 2 and header_value in image_columns.values():
                logger.info(f"[Image Match] Ảnh phù hợp tại dòng {img_row}, cột {header_value}")

                try:
                    logger.info(f"[Image File] Type of image.ref: {type(image.ref)}")
                    image.ref.seek(0)
                except Exception as seek_err:
                    error_msg = f"Lỗi seek ảnh tại cột {header_value}: {seek_err}"
                    logger.error(f"[Image Error] {error_msg}")
                    image_results[header_value] = (None, error_msg)
                    continue

                wrapped_file = SimpleNamespace(file=image.ref)

                try:
                    file_url = upload_file(wrapped_file, "images" if not is_primary else "images_primary")
                    logger.info(f"[Image Upload] Upload kết quả tại {header_value}: {file_url}")

                    if file_url:
                        if is_primary and header_value == "images_primary":
                            return file_url
                        image_results[header_value] = (ItemImageDBReq(
                            images_id=generate_id("IMAGE"),
                            images_url=file_url
                        ), None)
                    else:
                        image_results[header_value] = (None, f"Lỗi không upload được ảnh ở cột {header_value}")
                except Exception as upload_err:
                    error_msg = f"Lỗi upload ảnh tại cột {header_value}: {upload_err}"
                    logger.error(f"[Image Upload Error] {error_msg}")
                    image_results[header_value] = (None, error_msg)
        except Exception as e:
            logger.error(f"Lỗi xử lý ảnh tại row {img_row}: {e}")

    # Gán lỗi cho các cột không có ảnh
    for col_name in image_columns.values():
        if col_name not in image_results:
            image_results[col_name] = (None, f"Không tìm thấy ảnh ở cột {col_name}")

    if is_primary:
        url, error = image_results.get("images_primary", (None, None))
        if error:
            raise Exception(error)
        return url

    return image_results  # Dict[str, Tuple[Optional[ItemImageDBReq], Optional[str]]]

def extract_pdf_bytes_from_ole_bin(bin_stream: BytesIO) -> bytes:
    try:
        if olefile.isOleFile(bin_stream):
            bin_stream.seek(0)
            ole = olefile.OleFileIO(bin_stream)
            for stream_name in ole.listdir():
                data = ole.openstream(stream_name).read()
                if b"%PDF" in data:
                    start = data.find(b"%PDF")
                    pdf_data = data[start:]
                    ole.close()
                    return pdf_data
            ole.close()
        return None
    except Exception as e:
        print(f"OLE parse error: {e}")
        return None

async def extract_certificates_from_excel(file_stream: BytesIO, df: pd.DataFrame):
    zipf = zipfile.ZipFile(file_stream, "r")

    result_map = {}  # row_idx -> PDF URL

    for file_info in zipf.infolist():
        if file_info.filename.startswith("xl/embeddings/") and file_info.filename.endswith(".bin"):
            bin_data = zipf.read(file_info.filename)
            bin_stream = BytesIO(bin_data)

            # extract PDF bytes from .bin
            pdf_bytes = extract_pdf_bytes_from_ole_bin(bin_stream)
            if pdf_bytes:
                wrapped_file = SimpleNamespace(file=BytesIO(pdf_bytes))
                wrapped_file.content_type = "application/pdf"
                url = upload_any_file(wrapped_file, "certificates")

                # mapping theo thứ tự nhúng (ví dụ mapping theo row cùng index)
                row_idx = len(result_map)
                result_map[row_idx] = url
    return result_map

async def import_products(file: UploadFile, email: str):
    try:
        if redis.is_import_locked():
            raise response.JsonException(
                status_code=status.HTTP_400_BAD_REQUEST,
                message="Hệ thống đang nộp dữ liệu, vui lông truy cập sau"
            )

        redis.set_import_lock()

        if not file.filename.endswith(".xlsx"):
            raise response.JsonException(
                status_code=status.HTTP_400_BAD_REQUEST,
                message="Chỉ hỗ trợ file Excel .xlsx"
            )

        contents = await file.read()
        workbook = load_workbook(BytesIO(contents))
        sheet = workbook.active
        df = pd.read_excel(BytesIO(contents))
        df = df.fillna("")  # thay NaN bằng chuỗi rỗng

        error_messages = []
        image_columns = {
            0: "images_1", 1: "images_2", 2: "images_3",
            3: "images_4", 4: "images_5", 5: "images_6",
            6: "images_7", 7: "images_8"
        }

        categories = await get_all_categories_admin()

        certificate_map = await extract_certificates_from_excel(BytesIO(contents), df)

        collection = db[collection_name]
        inventory_collection = db[collection_inventory]

        for idx, row in df.iterrows():
            try:
                logger.info(f"Processing row: {idx}")
                excel_row_idx = idx + 1
                row_errors = []

                cur = collection.find_one({"slug": row.get("slug", "")})
                if cur:
                    error_messages.append(f"Dòng {excel_row_idx}: Sản phẩm slug {row.get('slug', '')}đã tồn tại")
                    continue

                # Extract certificate file
                certificate_url = certificate_map.get(idx, "")
                logger.info(f"Certificate url: {certificate_url}")

                # Extract images
                try:
                    image_result_map = await extract_images_direct(sheet, df, idx, image_columns)
                    image_list = []
                    for col, (img, err) in image_result_map.items():
                        if err:
                            error_messages.append(f"Dòng {excel_row_idx}: {err}")
                        elif img:
                            image_list.append(img)
                    if not image_list:
                        error_messages.append(f"Dòng {excel_row_idx}: Không có ảnh nào hợp lệ")
                except Exception as e:
                    row_errors.append(f"Dòng {excel_row_idx}: Lỗi hình ảnh - {e}")
                    image_list = []

                # Extract primary image
                try:
                    image_primary = await extract_images_direct(sheet, df, idx, {0: "images_primary"}, is_primary=True)
                    if not image_primary:
                        error_messages.append(f"Dòng {excel_row_idx}: Lỗi hình ảnh chính không xác định")
                except Exception as e:
                    row_errors.append(f"Dòng {excel_row_idx}: Lỗi ảnh chính - {e}")
                    image_primary = None

                product_id = generate_id("PRODUCT")

                # Extract prices
                try:
                    prices_raw = row.get("prices", "[]")
                    prices_list = json.loads(prices_raw)
                    prices = []
                    inventory_list = []
                    for price in prices_list:
                        price_obj = ItemPriceDBReq(
                                price_id=generate_id("PRICE"),
                                discount=price.get("discount", 0),
                                unit=price.get("unit", ""),
                                weight=price.get("weight", 0),
                                amount=price.get("amount", 0),
                                original_price=price.get("original_price", 0),
                                price=price.get("original_price", 0) * (100 - price.get("discount", 0)) / 100,
                                expired_date=price.get("expired_date", get_current_time()),
                                inventory=price.get("inventory", 0)
                            )
                        prices.append(price_obj)

                        inventory_obj = ItemProductInventoryReq(
                            product_id=product_id,
                            price_id=price_obj.price_id,
                            inventory=price_obj.inventory,
                            amount=price_obj.amount,
                        )
                        inventory_list.append(inventory_obj.dict())
                except Exception as e:
                    row_errors.append(f"Dòng {excel_row_idx}: Lỗi đọc prices - {e}")
                    prices = []
                    inventory_list = []

                # Extract category
                category_obj, cat_error = await extract_category_object(row, categories)
                if cat_error:
                    row_errors.append(f"Dòng {excel_row_idx} - {cat_error}")

                # Extract ingredients
                try:
                    ingredients_raw = row.get("ingredients", "[]")
                    ingredients_list = json.loads(ingredients_raw)
                    ingredients = [
                        ItemIngredientDBReq(
                            ingredient_name=ing.get("ingredient_name", ""),
                            ingredient_amount=ing.get("ingredient_amount", ""),
                        ) for ing in ingredients_list
                    ]
                except Exception as e:
                    row_errors.append(f"Dòng {excel_row_idx}: Lỗi đọc ingredients - {e}")
                    ingredients = []

                # Manufacturer
                try:
                    contact_str = normalize_manufacturer_contact(row.get("manufacture_contact", ""))
                    manufacturer = ItemManufacturerDBReq(
                        manufacture_name=row.get("manufacture_name", ""),
                        manufacture_address=row.get("manufacture_address", ""),
                        manufacture_contact=contact_str
                    )
                except Exception as e:
                    row_errors.append(f"Dòng {excel_row_idx}: Lỗi thông tin nhà sản xuất - {e}")
                    manufacturer = None

                if row_errors:
                    error_messages.extend(row_errors)
                    continue

                try:
                    product = ItemProductDBReq(
                        product_id=product_id,
                        product_name=row.get("product_name", ""),
                        name_primary=row.get("name_primary", ""),
                        prices=prices,
                        slug=row.get("slug", ""),
                        description=row.get("description", ""),
                        full_descriptions=row.get("full_descriptions", ""),
                        category=category_obj,
                        origin=row.get("origin", ""),
                        ingredients=ingredients,
                        uses=row.get("uses", ""),
                        dosage=row.get("dosage", ""),
                        side_effects=row.get("side_effects", ""),
                        precautions=row.get("precautions", ""),
                        storage=row.get("storage", ""),
                        manufacturer=manufacturer,
                        dosage_form=row.get("dosage_form", ""),
                        brand=row.get("brand", ""),
                        prescription_required=bool(row.get("prescription_required", False)),
                        registration_number=row.get("registration_number", ""),
                        images=image_list,
                        images_primary=image_primary,
                        certificate_file=certificate_url,
                        created_by=email,
                        updated_by=email
                    )
                    insert_result = collection.insert_one(product.dict())

                    logger.info(f"Thêm sản phẩm thành công: {insert_result.inserted_id}")

                    result = inventory_collection.insert_many(inventory_list)
                    logger.info(f"Đã thêm {len(inventory_list)} bản ghi vào bảng products_inventory")
                except ValidationError as e:
                    error_messages.append(f"Dòng {excel_row_idx}: Lỗi dữ liệu không hợp lệ - {e}")

            except Exception as e:
                error_messages.append(f"Dòng {idx + 1}: Lỗi không xác định - {e}")

        file.file.seek(0)
        excel_url = upload_any_file(file, folder="import_products")
        logger.info(f"Uploaded original Excel file to S3: {excel_url}")

        import_item = ItemProductImportReq(
            import_id=generate_id("IMPORT"),
            file_url=excel_url,
            error_message=error_messages
        )
        db[collection_import].insert_one(import_item.dict())
        logger.info(f"Import log written to '{collection_import}'")

    except Exception as e:
        logger.error(f"Error importing product: {str(e)}")
        raise e
    finally:
        redis.clear_import_lock()

async def get_imported_products(page: int, page_size: int):
    try:
        collection = db[collection_import]
        skip_count = (page - 1) * page_size
        imports_list = list(collection.find({}, {"_id": 0}).sort("_id", -1).skip(skip_count).limit(page_size))
        total = collection.count_documents({})
        return {
            "total_imports": total,
            "imports": [ItemProductImportRes(**imports) for imports in imports_list]
        }
    except Exception as e:
        logger.error(f"Error getting imported products: {str(e)}")
        raise e

async def delete_imported_products(import_id: str):
    try:
        collection = db[collection_import]
        import_item = collection.find_one({"import_id": import_id})
        if not import_item:
            raise response.JsonException(status_code=status.HTTP_400_BAD_REQUEST, message="Không tìm thấy Import")
        else:
            delete_result = collection.delete_one({"import_id": import_id})
            if delete_result.deleted_count == 0:
                logger.info(f"Import not deleted for import_id: {import_id}")
            return response.SuccessResponse(message="Xóa Import thành công")
    except Exception as e:
        logger.error(f"Error deleting imported product: {str(e)}")
        raise e

async def check_all_product_discount_expired():
    try:
        collection = db[collection_name]
        now = get_current_time()
        products = collection.find({
            "prices": {
                "$elemMatch": {
                    "discount": {"$gt": 0},
                    "expired_date": {"$lt": now}
                }
            }
        })

        updates = []

        for product in products:
            updated_prices = []
            changed = False

            for price in product.get("prices", []):
                if price.get("discount", 0) > 0 and price.get("expired_date") and price["expired_date"] < now:
                    price["discount"] = 0
                    price["price"] = price.get("original_price", 0)
                    changed = True
                updated_prices.append(price)

            if changed:
                updates.append(
                    UpdateOne(
                        {"product_id": product["product_id"]},
                        {"$set": {"prices": updated_prices}}
                    )
                )

        if updates:
            result = collection.bulk_write(updates)
            logger.info(f"[discount_expired] Matched: {result.matched_count}, Modified: {result.modified_count}")
            return result.modified_count
        else:
            logger.info("No expired discount to update.")
            return 0
    except Exception as e:
        logger.error(f"Failed to update expired discounts: {e}")
        raise e

async def normalize_products_inventory():
    product_collection = db[collection_name]
    inventory_collection = db[collection_inventory]

    products = list(product_collection.find({}, {"product_id": 1, "prices": 1}))
    bulk_ops = []

    for product in products:
        product_id = product["product_id"]
        for price in product.get("prices", []):
            price_id = price.get("price_id")
            if not price_id:
                continue

            # Tạo bản ghi inventory từ toàn bộ thông tin trong price
            inventory_doc = {
                "product_id": product_id,
                "price_id": price_id,
                "amount": price.get("amount", 0),
                "inventory": price.get("inventory", 0),
                "sell": price.get("sell", 0),
                "delivery": price.get("delivery", 0)
            }

            bulk_ops.append(ReplaceOne(
                {"product_id": product_id, "price_id": price_id},
                inventory_doc,
                upsert=True
            ))

    if bulk_ops:
        inventory_collection.bulk_write(bulk_ops)

    print(f"Đã cập nhật đầy đủ {len(bulk_ops)} bản ghi từ prices sang products_inventory.")

async def get_low_stock_products():
    try:
        collection = db[collection_name]

        pipeline = [
            {"$unwind": "$prices"},
            {"$match": {"$expr": {"$lte": [{"$subtract": ["$prices.inventory", "$prices.sell"]}, 10]}}},
            {
                "$project": {
                    "_id": 0,
                    "product_id": 1,
                    "product_name": 1,
                    "images_primary": 1,
                    "prices_id": "$prices.price_id",
                    "unit": "$prices.unit",
                    "sell": "$prices.sell",
                    "delivery": "$prices.delivery",
                    "inventory": "$prices.inventory",
                }
            }
        ]

        cursor = collection.aggregate(pipeline)
        result = []
        for doc in cursor:
            result.append(doc)

        return result
    except Exception as e:
        logger.error(f"Failed [get_low_stock_products]: {e}")
        raise e